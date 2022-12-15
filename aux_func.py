# -*- coding: utf-8 -*-

def dir_list(path):
    import os
    dir_list = []

    for dirpath, dirnames, filenames in os.walk(os.path.normpath(path)):
        for name in filenames:
            file = os.path.join(dirpath, name)
            if file.endswith('.pdf'):
                dir_list.append(file)
    return dir_list


def doc_list(list):
    doc_list = []
    for filename in list:
        doc_list.append(filename)
    new_list = set(doc_list)
    return new_list


def merger(path, doclist):
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side
    wb = load_workbook(path)

    ws = wb.get_sheet_by_name('Summary')
    adress_dict = {}

    for i in doclist:
        adress_list = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == i:
                    adress_list.append(cell.coordinate)
        adress_dict[i] = adress_list

    for value in adress_dict.values():
        adress = '{0}{1}{2}'.format(value[0], ':', value[len(value) - 1])
        ws.merge_cells(adress)

    for row in ws.iter_rows():
        for cell in row:
            thins = Side(border_style='thin', color='000000')
            cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

    wb.save(path)


def styler(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side
    wb = load_workbook(path)

    for i in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(i)
        for row in ws.iter_rows():
            for cell in row:
                thins = Side(border_style='thin', color='000000')
                cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

    wb.save(path)


def write(outpath, dict_to_write):
    import pandas as pd

    with pd.ExcelWriter(path=outpath, engine='xlsxwriter') as wb:
        for key, value in dict_to_write.items():
            frame = value
            sheet_name = key
            frame.to_excel(wb, sheet_name=sheet_name, index=False)
            sheet = wb.sheets[sheet_name]

            for column in frame:
                base_column_width = max(frame[column].astype(str).map(len).max(), len(column)) + 4
                column_width = 1.05 * base_column_width
                col_idx = frame.columns.get_loc(column)

                format = wb.book.add_format()
                format.set_align('center')
                format.set_align('vcenter')

                sheet.set_column(col_idx, col_idx, column_width, format)
    styler(outpath)


def frame_list_to_sict(list_to_write):
    dict_to_write = {}

    for i in range(len(list_to_write)):
        dict_to_write[list_to_write[i].name] = list_to_write[i].out

    return dict_to_write


def create_frame_list(dir_list):
    import classes as cl
    frame_list = []
    for file in dir_list:
        new_frame = cl.PDFscan(path=file)
        new_frame.watch()
        frame_list.append(new_frame.base_frame())

    return frame_list


